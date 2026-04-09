"""
Motor de matching y generación de pólizas CONTPAq.
Versión 3.0 — GBC Business Consulting / RZ2 Sistemas

Nuevo paradigma v3.0:
  El archivo REP es la FUENTE PRIMARIA de matching.
  Cada UUID_REP agrupa las facturas que ampara y define el monto del depósito.
  El estado de cuenta bancario se usa como VALIDACIÓN (monto ± tolerancia, fecha ± 2 días).
  Fact_Pendientes enriquece con IVA, subtotal, cuenta CXC e IVA_CONSIDERADO.
  Depósitos sin REP → Sin Match - Revisar.

Flujo:
  1. Leer REP  → agrupar por UUID_REP → calcular total por grupo
  2. Leer banco → solo abonos (+)
  3. Matching  → cada grupo REP busca su abono en banco (monto ± tol, fecha ± 2d)
  4. Enriquecer → buscar cada UUID_factura en Fact_Pendientes
  5. Generar   → bloque CONTPAq con estructura correcta
  6. Sin match → abonos bancarios que ningún REP reclamó
"""

import pandas as pd
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

# ── Cuentas fijas ──────────────────────────────────────────────────────────────
CTA_BANCO     = 10201001
CTA_IVA_PXAC  = 20901000
CTA_IVA_TRAS  = 20801000
TIPO_POLIZA   = 11
TOLERANCIA    = 0.10   # diferencia máxima de monto permitida
DIAS_MARGEN   = 2      # diferencia máxima de días entre REP y banco


# ══════════════════════════════════════════════════════════════════════════════
# LECTURA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════════════════

def leer_banco(ruta):
    """Lee el CSV de Scotiabank y retorna solo los abonos (+)."""
    df = pd.read_csv(ruta, encoding='latin-1')
    for col in df.select_dtypes(include='object').columns:
        df[col] = df[col].astype(str).str.strip().str.strip("'")
    df = df[df['Cargo/Abono'].str.strip() == '+'].copy()
    df['Fecha_dt'] = pd.to_datetime(df['Fecha'], format='%d%m%Y', errors='coerce')
    df['Importe']  = pd.to_numeric(df['Importe'], errors='coerce').fillna(0)
    return df.reset_index(drop=True)


def leer_reps(ruta):
    """
    Lee el archivo REP y agrupa por UUID_REP (col D, idx 3).

    Estructura por fila:
      - Cada fila = una factura siendo pagada
      - UUID_REP (D/3)      : UUID del complemento — se repite si cubre varios folios
      - Fecha pago (AI/34)  : fecha del depósito bancario
      - Monto (AV/47)       : monto TOTAL del depósito — constante dentro del grupo
      - UUID_fac (AW/48)    : UUID de la factura que se cubre
      - Folio doc (BC/54)   : folio de la factura
      - Importe pagado (BW/74): monto cobrado de ESTA factura (para asientos CXC)

    Retorna lista de grupos:
      uuid_rep   : str
      fecha_pago : datetime
      rfc        : str
      razon      : str
      monto_dep  : float  ← AV, monto total del depósito (para cruzar con banco)
      facturas   : list de {uuid_fac, folio, importe_fac}
                             importe_fac = BW, lo que se cobra de cada factura
    """
    wb   = load_workbook(ruta, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    grupos = defaultdict(list)
    meta   = {}

    for r in rows[1:]:
        try:
            estado = str(r[17]).upper() if r[17] else ''
            if 'VIGENTE' not in estado:
                continue
            uuid_rep    = str(r[3]).strip()  if r[3]  else ''
            uuid_fac    = str(r[48]).strip() if r[48] else ''
            folio       = str(r[54]).strip() if r[54] else ''
            monto_dep   = float(r[47])       if r[47] else 0.0  # AV: total depósito
            importe_fac = float(r[74])       if r[74] else 0.0  # BW: cobrado esta factura
            rfc         = str(r[12]).strip().upper() if r[12] else ''
            razon       = str(r[13]).strip().upper() if r[13] else ''
            fecha_raw   = r[34]

            if not uuid_rep or not uuid_fac:
                continue

            if isinstance(fecha_raw, datetime):
                fecha_dt = fecha_raw
            else:
                fecha_dt = pd.to_datetime(str(fecha_raw), errors='coerce')
                fecha_dt = None if pd.isna(fecha_dt) else fecha_dt.to_pydatetime()

            grupos[uuid_rep].append({
                'uuid_fac':    uuid_fac,
                'folio':       folio,
                'importe_fac': importe_fac,  # BW: monto del asiento CXC
            })
            if uuid_rep not in meta:
                meta[uuid_rep] = {
                    'monto_dep': monto_dep,  # AV: monto del depósito bancario
                    'fecha':     fecha_dt,
                    'rfc':       rfc,
                    'razon':     razon,
                }
        except Exception:
            continue

    resultado = []
    for uuid_rep, facs in grupos.items():
        m = meta[uuid_rep]
        resultado.append({
            'uuid_rep':   uuid_rep,
            'fecha_pago': m['fecha'],
            'rfc':        m['rfc'],
            'razon':      m['razon'],
            'monto_dep':  round(m['monto_dep'], 2),  # AV → cruzar con banco
            'facturas':   facs,
        })
    return resultado


def leer_facturas(ruta):
    """
    Lee Fact_Pendientes.
    Retorna dict: uuid_lower → row con campos normalizados.
    """
    df = pd.read_excel(ruta, sheet_name='Fact_Pendientes', header=0)

    col_map = {
        df.columns[6]:  'UUID',
        df.columns[9]:  'Serie',
        df.columns[10]: 'Folio',
        df.columns[11]: 'Tipo',
        df.columns[17]: 'RFC_receptor',
        df.columns[18]: 'Razon_receptor',
        df.columns[26]: 'Estado',
        df.columns[29]: 'Metodo_pago',
        df.columns[30]: 'Forma_pago',
        df.columns[31]: 'SubTotal',
        df.columns[33]: 'IVA',
        df.columns[36]: 'Total',
        df.columns[37]: 'IVA_CONSIDERADO',
    }
    df = df.rename(columns=col_map)
    df = df[df['Tipo'].astype(str).str.contains('Ingreso', na=False)]
    df = df[df['Estado'].astype(str).str.upper().str.contains('VIGENTE', na=False)]

    df['RFC_limpio']       = df['RFC_receptor'].astype(str).str.strip().str.upper()
    df['Folio_str']        = df['Folio'].astype(str).str.strip()
    df['Total_num']        = pd.to_numeric(df['Total'], errors='coerce').fillna(0)
    df['IVA_num']          = pd.to_numeric(df['IVA'],   errors='coerce').fillna(0)
    df['IVA_ya_declarado'] = df['IVA_CONSIDERADO'].astype(str).str.strip().str.upper() == 'SI'
    df['Metodo_pago']      = df['Metodo_pago'].astype(str)

    df = df.dropna(subset=['UUID']).reset_index(drop=True)

    # Índice por UUID (para enriquecer desde REP)
    fact_map = {str(row['UUID']).strip().lower(): row for _, row in df.iterrows()}
    # DataFrame completo (para Pasada 2 — búsqueda por folio/RFC/monto)
    return fact_map, df


def leer_catalogo(ruta):
    """
    Soporta dos formatos:
      A) Hoja 'Cat_clientes': Nombre | RFC | CodigoCuenta  (nuevo)
      B) Hoja 'cuentas': posicional col1=num col2=nombre col3=rfc (legado)
    """
    wb  = load_workbook(ruta, read_only=True, data_only=True)
    cat = {}

    if 'Cat_clientes' in wb.sheetnames:
        ws   = wb['Cat_clientes']
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            nom, rfc, num = row[0], row[1], row[2]
            if not num:
                continue
            num_str = str(int(num)) if isinstance(num, (int, float)) else str(num).strip()
            cat[num_str] = {
                'nombre': str(nom).strip().upper() if nom else '',
                'rfc':    str(rfc).strip().upper() if rfc else '',
            }
        return cat

    ws = wb['cuentas']
    for row in ws.iter_rows(values_only=True):
        num, nom = row[1], row[2]
        rfc = row[3] if len(row) > 3 else None
        if num and str(num).startswith('1050') and len(str(num)) >= 8:
            cat[str(int(num))] = {
                'nombre': str(nom).strip().upper() if nom else '',
                'rfc':    str(rfc).strip().upper() if rfc else '',
            }
    return cat


# ══════════════════════════════════════════════════════════════════════════════
# MATCHING — nuevo paradigma REP como fuente primaria
# ══════════════════════════════════════════════════════════════════════════════

def hacer_matching(grupos_rep, banco):
    """
    Para cada grupo REP busca el abono bancario correspondiente.
    Criterio: monto total del grupo ≈ importe banco (±TOLERANCIA)
              Y fecha_pago REP ≈ fecha banco (±DIAS_MARGEN días)

    Retorna:
      matches   → lista de dicts: {grupo_rep, banco_row}
      sin_rep   → abonos bancarios que ningún REP reclamó
    """
    banco_disponible = banco.copy()
    matches  = []
    sin_match_rep = []   # grupos REP sin abono bancario (raro, pero posible)

    for grupo in grupos_rep:
        total      = grupo['monto_dep']   # AV: monto real del depósito
        fecha_rep  = grupo['fecha_pago']
        encontrado = False

        # Filtrar banco por monto
        cands = banco_disponible[
            abs(banco_disponible['Importe'] - total) <= TOLERANCIA
        ]

        # Si hay fecha en el REP, filtrar por ventana de días
        if fecha_rep is not None and not cands.empty:
            fecha_rep_dt = pd.Timestamp(fecha_rep)
            ventana = cands[
                abs((cands['Fecha_dt'] - fecha_rep_dt).dt.days) <= DIAS_MARGEN
            ]
            if not ventana.empty:
                cands = ventana

        if not cands.empty:
            # Tomar el primer candidato (monto + fecha más cercana)
            idx = cands.index[0]
            matches.append({
                'grupo_rep': grupo,
                'banco_row': banco_disponible.loc[idx],
            })
            banco_disponible = banco_disponible.drop(idx)
            encontrado = True

        if not encontrado:
            sin_match_rep.append(grupo)

    # Lo que quedó en banco sin ser reclamado por ningún REP
    sin_rep = [banco_disponible.loc[i] for i in banco_disponible.index]

    return matches, sin_match_rep, sin_rep


def extraer_folios(texto):
    """Extrae números de 3-5 dígitos del concepto bancario (posibles folios)."""
    nums = re.findall(r'\b(\d{3,5})\b', texto or '')
    return [n for n in nums if 100 <= int(n) <= 99999]


def matching_pasada2(sin_rep_banco, fact_df, uuids_usados_p1):
    """
    Pasada 2: depósitos bancarios sin REP → buscar en Fact_Pendientes.

    Candidatas: PUE vigentes + PPD vigentes sin REP (no usadas en Pasada 1).
    Jerarquía:
      1. Folio en concepto banco + RFC ordenante
      2. Monto depósito ≈ Total factura (±TOLERANCIA) + RFC ordenante
      3. RFC ordenante único en Fact_Pendientes

    Retorna:
      matches_p2  → lista de dicts compatibles con escribir_excel
      sin_match_p2 → depósitos que tampoco matchearon en Pasada 2
    """
    # Filtrar candidatas: PUE + PPD, excluir las ya usadas en Pasada 1
    candidatas = fact_df[
        ~fact_df['UUID'].astype(str).str.strip().str.lower().isin(uuids_usados_p1)
    ].copy()

    matches_p2   = []
    sin_match_p2 = []
    uuids_usados = set(uuids_usados_p1)

    for _, dep in (pd.DataFrame(sin_rep_banco) if not isinstance(sin_rep_banco, pd.DataFrame)
                   else sin_rep_banco).iterrows():
        rfc_dep     = str(dep.get('RFC Ordenante', '')).strip().upper()
        monto_dep   = float(dep.get('Importe', 0))
        concepto    = str(dep.get('Concepto_full', dep.get('Concepto', ''))).upper()
        folios_dep  = extraer_folios(concepto)

        # Pool de candidatas no usadas aún en esta pasada
        pool = candidatas[
            ~candidatas['UUID'].astype(str).str.strip().str.lower().isin(uuids_usados)
        ]

        matched_fac = None
        metodo      = None

        # 1. Folio en concepto + RFC
        if folios_dep and rfc_dep:
            for folio in folios_dep:
                cands = pool[pool['Folio_str'] == folio]
                por_rfc = cands[cands['RFC_limpio'] == rfc_dep]
                if not por_rfc.empty:
                    matched_fac = por_rfc.iloc[0]
                    metodo = 'p2:folio+rfc'
                    break
            if matched_fac is None:
                # Folio sin RFC conocido
                for folio in folios_dep:
                    cands = pool[pool['Folio_str'] == folio]
                    if len(cands) == 1:
                        matched_fac = cands.iloc[0]
                        metodo = 'p2:folio'
                        break

        # 2. Monto exacto + RFC
        if matched_fac is None:
            cands = pool[abs(pool['Total_num'] - monto_dep) <= TOLERANCIA]
            if rfc_dep:
                por_rfc = cands[cands['RFC_limpio'] == rfc_dep]
                if not por_rfc.empty:
                    matched_fac = por_rfc.iloc[0]
                    metodo = 'p2:monto+rfc'
            if matched_fac is None and len(cands) == 1:
                matched_fac = cands.iloc[0]
                metodo = 'p2:monto_exacto'

        # 3. RFC único en el pool
        if matched_fac is None and rfc_dep:
            cands = pool[pool['RFC_limpio'] == rfc_dep]
            if len(cands) == 1:
                matched_fac = cands.iloc[0]
                metodo = 'p2:rfc_unico'

        if matched_fac is not None:
            uuid_fac = str(matched_fac['UUID']).strip().lower()
            uuids_usados.add(uuid_fac)
            # Construir un "grupo_rep sintético" compatible con generar_bloque
            grupo_sintetico = {
                'uuid_rep':   None,   # sin REP
                'razon':      str(matched_fac.get('Razon_receptor', '')).upper().strip(),
                'rfc':        str(matched_fac.get('RFC_limpio', '')),
                'monto_dep':  monto_dep,
                'fecha_pago': dep.get('Fecha_dt'),
                'facturas': [{
                    'uuid_fac':    str(matched_fac['UUID']).strip(),
                    'folio':       str(matched_fac['Folio_str']),
                    'importe_fac': round(float(matched_fac['Total_num']), 2),
                }],
            }
            matches_p2.append({
                'grupo_rep': grupo_sintetico,
                'banco_row': dep,
                'metodo':    metodo,
                'pasada':    2,
            })
        else:
            sin_match_p2.append(dep)

    return matches_p2, sin_match_p2


# ══════════════════════════════════════════════════════════════════════════════
# BÚSQUEDA DE CUENTA CXC
# ══════════════════════════════════════════════════════════════════════════════

def buscar_cuenta_cte(rfc, razon, catalogo):
    """Jerarquía: 1° RFC exacto  2° Nombre exacto  3° Palabras clave"""
    rfc   = str(rfc).upper().strip()
    razon = str(razon).upper().strip()

    if rfc:
        for num, datos in catalogo.items():
            if datos['rfc'] == rfc:
                return num

    for num, datos in catalogo.items():
        if datos['nombre'] == razon:
            return num

    palabras = [p for p in razon.split() if len(p) > 3]
    for num, datos in catalogo.items():
        if palabras and all(p in datos['nombre'] for p in palabras):
            return num

    return 'SIN_CUENTA'


# ══════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DE BLOQUES CONTPAq
# ══════════════════════════════════════════════════════════════════════════════

def generar_bloque(match, num_pol, fact_map, catalogo):
    """
    Genera las filas CONTPAq para un match REP ↔ banco.

    Pago simple (1 factura en el REP):
        P   YYYYMMDD  num_pol
        M1  BANCO      folio   D  total
        M1  IVA_PXAC   folio   D  iva      (omitir si iva=0)
        M1  IVA_TRAS   folio   H  iva      (omitir si iva=0)
        M1  CXC        folio   H  total
        AD  UUID_FACTURA
        AD  UUID_REP

    Multi-folio (N facturas en el REP):
        P   YYYYMMDD  num_pol
        M1  BANCO      VARIAS  D  total_suma
        — por cada factura:
        M1  IVA_PXAC   folio_n  D  iva_n    (omitir si iva_n=0)
        M1  CXC        folio_n  H  total_n
        — acumulado al final:
        M1  IVA_TRAS   VARIAS   H  iva_total (omitir si iva_total=0)
        AD  UUID_FACTURA_1
        AD  UUID_FACTURA_2 ...
        AD  UUID_REP
    """
    grupo    = match['grupo_rep']
    banco    = match['banco_row']
    uuid_rep = grupo['uuid_rep']
    es_multi = len(grupo['facturas']) > 1

    fecha_banco = banco['Fecha_dt']
    # CONTPAq requiere fecha real (datetime) con formato aaaammdd — NO entero
    fecha_dt = fecha_banco.to_pydatetime() if pd.notna(fecha_banco) else datetime(2000, 1, 1)
    total_banco = round(float(banco['Importe']), 2)

    if es_multi:
        desc_pol = f"COBRANZA {grupo['razon']} - {len(grupo['facturas'])} FACTURAS"
    else:
        desc_pol = f"COBRANZA {grupo['razon']}"
    desc_pol = desc_pol[:80]

    filas = []
    # P: fecha como datetime, TipoPol=1, D/C como enteros
    filas.append(('P', fecha_dt, 1, num_pol, 1, '0', desc_pol, TIPO_POLIZA, 0, 0))

    avisos = []

    def _folio_int(folio_str):
        """Convierte folio a entero si es puramente numérico, si no lo deja como str."""
        try:
            return int(folio_str)
        except (ValueError, TypeError):
            return folio_str

    def _cta_int(cta):
        """Convierte cuenta a entero."""
        try:
            return int(cta)
        except (ValueError, TypeError):
            return cta

    if not es_multi:
        # ── Pago simple ───────────────────────────────────────────────────────
        fac_rep  = grupo['facturas'][0]
        uuid_fac = fac_rep['uuid_fac'].lower()
        folio    = _folio_int(fac_rep['folio'])
        importe_fac = round(fac_rep['importe_fac'], 2)
        fac_data = fact_map.get(uuid_fac)

        if fac_data is not None:
            total_fac = round(float(fac_data['Total_num']), 2)
            iva_fac   = round(float(fac_data['IVA_num']), 2)
            if fac_data['IVA_ya_declarado'] or total_fac == 0:
                iva = 0
            else:
                proporcion = importe_fac / total_fac if total_fac else 1
                iva = round(iva_fac * proporcion, 2)
            cta  = _cta_int(buscar_cuenta_cte(fac_data['RFC_limpio'], fac_data['Razon_receptor'], catalogo))
            desc = f"COBRANZA {str(fac_data['Razon_receptor']).upper().strip()}"[:80]
        else:
            iva  = 0
            cta  = _cta_int(buscar_cuenta_cte(grupo['rfc'], grupo['razon'], catalogo))
            desc = desc_pol
            avisos.append(f'UUID {fac_rep["uuid_fac"][:8]}... no encontrado en Fact_Pendientes')

        # TipoMovto: 0=Debe, 1=Haber (enteros, como CONTPAq espera)
        filas.append(('M1', CTA_BANCO,    folio, 0, total_banco, '0', 0, desc))
        if iva > 0:
            filas.append(('M1', CTA_IVA_PXAC, folio, 0, iva,         '0', 0, desc))
            filas.append(('M1', CTA_IVA_TRAS, folio, 1, iva,         '0', 0, desc))
        filas.append(    ('M1', cta,           folio, 1, importe_fac, '0', 0, desc))
        filas.append(('AD', fac_rep['uuid_fac']))
        if uuid_rep:   # Pasada 2 no tiene REP
            filas.append(('AD', uuid_rep))

    else:
        # ── Multi-folio ───────────────────────────────────────────────────────
        filas.append(('M1', CTA_BANCO, 'VARIAS', 0, total_banco, '0', 0, desc_pol))

        iva_total  = 0.0
        uuids_facs = []

        for fac_rep in grupo['facturas']:
            uuid_fac    = fac_rep['uuid_fac'].lower()
            folio       = _folio_int(fac_rep['folio'])
            importe_fac = round(fac_rep['importe_fac'], 2)
            fac_data    = fact_map.get(uuid_fac)

            if fac_data is not None:
                total_fac = round(float(fac_data['Total_num']), 2)
                iva_fac   = round(float(fac_data['IVA_num']), 2)
                if fac_data['IVA_ya_declarado'] or total_fac == 0:
                    iva = 0
                else:
                    proporcion = importe_fac / total_fac if total_fac else 1
                    iva = round(iva_fac * proporcion, 2)
                cta  = _cta_int(buscar_cuenta_cte(fac_data['RFC_limpio'], fac_data['Razon_receptor'], catalogo))
                desc = f"COBRANZA {str(fac_data['Razon_receptor']).upper().strip()}"[:80]
            else:
                iva  = 0
                cta  = _cta_int(buscar_cuenta_cte(grupo['rfc'], grupo['razon'], catalogo))
                desc = desc_pol
                avisos.append(f'UUID {fac_rep["uuid_fac"][:8]}... no encontrado en Fact_Pendientes')

            if iva > 0:
                filas.append(('M1', CTA_IVA_PXAC, folio, 0, iva,         '0', 0, desc))
            filas.append(    ('M1', cta,           folio, 1, importe_fac, '0', 0, desc))
            iva_total += iva
            uuids_facs.append(fac_rep['uuid_fac'])

        iva_total = round(iva_total, 2)
        if iva_total > 0:
            filas.append(('M1', CTA_IVA_TRAS, 'VARIAS', 1, iva_total, '0', 0, desc_pol))

        for uuid_f in uuids_facs:
            filas.append(('AD', uuid_f))
        if uuid_rep:   # Pasada 2 no tiene REP
            filas.append(('AD', uuid_rep))

    return filas, avisos


# ══════════════════════════════════════════════════════════════════════════════
# ESCRITURA DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def escribir_excel(matches, sin_match_rep, sin_rep_banco, fact_map, catalogo, ruta_out,
                   matches_p2=None, sin_match_p2=None):
    matches_p2   = matches_p2   or []
    sin_match_p2 = sin_match_p2 or []
    todos_matches = matches + matches_p2

    wb = Workbook()
    # Layout CONTPAq va primero — CONTPAq solo lee la primera hoja
    ws_lay = wb.active
    ws_lay.title = 'Layout CONTPAq'

    # Paleta
    fill_header = PatternFill('solid', start_color='1F3864')
    fill_P      = PatternFill('solid', start_color='D9E1F2')
    fill_P2     = PatternFill('solid', start_color='E8F5E9')   # verde claro — Pasada 2
    fill_AD     = PatternFill('solid', start_color='E2EFDA')
    fill_sm     = PatternFill('solid', start_color='FCE4D6')
    fill_smrep  = PatternFill('solid', start_color='EDE7F6')
    fill_dash   = PatternFill('solid', start_color='EBF3FB')
    fill_warn   = PatternFill('solid', start_color='FFF2CC')
    fnt_hdr     = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
    fnt_P       = Font(bold=True, name='Calibri', size=9)
    fnt_n       = Font(name='Calibri', size=9)
    fnt_title   = Font(bold=True, name='Calibri', size=14, color='1F3864')
    fnt_section = Font(bold=True, name='Calibri', size=11, color='1F3864')

    def hdr_row(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = fill_header
            cell.font = fnt_hdr
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22

    # Estadísticas Pasada 1
    n_ok     = len(matches)
    n_multi  = sum(1 for m in matches if len(m['grupo_rep']['facturas']) > 1)
    n_sm_rep = len(sin_match_rep)
    n_sm_ban = len(sin_rep_banco)   # entraron a Pasada 2

    # Estadísticas Pasada 2
    n_ok_p2  = len(matches_p2)
    n_sm_p2  = len(sin_match_p2)   # sin match en ninguna pasada

    # Facturas con IVA en cero (ambas pasadas)
    n_iva_cero = 0
    for m in todos_matches:
        for fac_rep in m['grupo_rep']['facturas']:
            fd = fact_map.get(fac_rep['uuid_fac'].lower())
            if fd is not None and fd['IVA_ya_declarado']:
                n_iva_cero += 1

    total_ok   = sum(float(m['banco_row']['Importe']) for m in matches)
    total_ok_p2= sum(float(m['banco_row']['Importe']) for m in matches_p2)
    total_sm   = sum(float(r.get('Importe', 0)) for r in sin_match_p2)
    total_procesos = n_ok + n_sm_ban
    tasa_p1 = f"{n_ok/total_procesos*100:.1f}%" if total_procesos else "0%"
    tasa_global = f"{(n_ok+n_ok_p2)/total_procesos*100:.1f}%" if total_procesos else "0%"

    # ── Dashboard ─────────────────────────────────────────────────────────────
    ws_dash = wb.create_sheet('Dashboard')
    dash = [
        ('SISTEMA DE PÓLIZAS DE INGRESO',          None),
        ('RZ2 Sistemas — GBC Business Consulting',  None),
        (None, None),
        ('RESULTADO DEL PROCESAMIENTO',             None),
        ('Abonos procesados',                       total_procesos),
        (None, None),
        ('── PASADA 1 (REP como fuente primaria)',  None),
        ('✅  Pólizas con REP',                     n_ok),
        ('   — Pago simple',                        n_ok - n_multi),
        ('   — Multi-folio',                        n_multi),
        ('   Tasa de match Pasada 1',               tasa_p1),
        ('⚠️  REPs sin abono bancario',             n_sm_rep),
        (None, None),
        ('── PASADA 2 (banco vs Fact_Pendientes)',   None),
        ('✅  Pólizas sin REP (PUE/PPD)',           n_ok_p2),
        ('⚠️  Sin match en ambas pasadas',          n_sm_p2),
        (None, None),
        ('TOTAL pólizas generadas',                 n_ok + n_ok_p2),
        ('Tasa global de match',                    tasa_global),
        (None, None),
        ('Monto pólizas P1',                        round(total_ok, 2)),
        ('Monto pólizas P2',                        round(total_ok_p2, 2)),
        ('Monto sin match',                         round(total_sm, 2)),
        (None, None),
        ('ℹ️  Facturas con IVA en cero',            n_iva_cero),
        (None, None),
        ('Fecha de proceso',                        datetime.now().strftime('%d/%m/%Y %H:%M')),
    ]
    for r, (lbl, val) in enumerate(dash, 1):
        cell_l = ws_dash.cell(row=r, column=1, value=lbl)
        cell_v = ws_dash.cell(row=r, column=2, value=val)
        cell_l.font = fnt_title if r == 1 else (fnt_section if r == 4 else fnt_n)
        # Filas con datos numéricos — fondo azul claro
        if lbl and lbl not in (None,) and val is not None and r > 4:
            cell_l.fill = fill_dash
            cell_v.fill = fill_dash
            cell_v.font = Font(bold=True, name='Calibri', size=10)
        # Filas de advertencia (sin match, REPs sin abono)
        if lbl and ('⚠' in str(lbl)) and val:
            cell_l.fill = fill_warn
            cell_v.fill = fill_warn
        if isinstance(val, float):
            cell_v.number_format = '$#,##0.00'
    ws_dash.column_dimensions['A'].width = 34
    ws_dash.column_dimensions['B'].width = 22

    # ── Layout CONTPAq ────────────────────────────────────────────────────────
    # (ws_lay ya fue creada como primera hoja al inicio de escribir_excel)

    # 16 filas de encabezado
    ENCABEZADO_CONTPAQ = [
        ['Asociación documento(AD)', 'UUID'],
        ['Devolución de IVA (IETU)(W)', 'IETUDeducible', 'SEPTIEMBRE 2019'],
        ['Devolución de IVA(V)', 'IdProveedor', 'ImpTotal', 'PorIVA', 'ImpBase', 'ImpIVA', 'CausaIVA', 'ExentoIVA', 'Serie', 'Folio', 'Referencia', 'OtrosImptos', 'ImpSinRet', 'IVARetenido', 'ISRRetenido', 'GranTotal', 'EjercicioAsignado', 'PeriodoAsignado', 'IdCuenta', 'IVAPagNoAcred', 'UUID'],
        ['Causación de IVA (Concepto de IETU)(E)', 'IdConceptoIETU'],
        ['Comprobantes(MC)', 'IdCuentaFlujoEfectivo', 'IdSegmentoNegCtaFlujo', 'Fecha', 'Serie', 'Folio', 'UUID', 'ClaveRastreo', 'Referencia', 'IdProveedor', 'CodigoConceptoIETU', 'ImpNeto', 'ImpNetoME', 'IdCuentaNeto', 'IdSegmentoNegNeto', 'PorIVA', 'ImporteIVA', 'ImporteIVAME', 'IVATasaExcenta', 'IdCuentaIVA', 'IdSegmentoNegIVA', 'NombreImpuesto', 'ImpImpuesto', 'ImpImpuestoME', 'IdCuentaImpuesto', 'IdSegmentoNegImp', 'ImpOtrosGastos', 'ImpOtrosGastosME', 'IdCuentaOtrosGastos', 'IdSegmentoNegOtrosGastos', 'IVARetenido', 'IVARetenidoME', 'IdCuentaRetIVA', 'IdSegmentoNegRetIVA', 'ISRRetenido', 'ISRRetenidoME', 'IdCuentaRetISR', 'IdSegmentoNegRetISR', 'NombreOtrasRetenciones', 'ImpOtrasRetenciones', 'ImpOtrasRetencionesME', 'IdCuentaOtrasRetenciones', 'IdSegmentoNegOtrasRet', 'BaseIVADIOT', 'BaseIETU', 'IVANoAcreditable', 'ImpTotalErogacion', 'IVAAcreditable', 'ImpExtra1', 'ImpExtra2', 'IdCategoria', 'IdSubCategoria', 'TipoCambio', 'IdDocGastos', 'EsCapturaCompleta', 'FolioStr'],
        ['Causación de IVA (IETU)(D)', 'IVATasa15NoAcred', 'IVATasa10NoAcred', 'IETU', 'Modificado', 'Origen', 'TotTasa16', 'BaseTasa16', 'IVATasa16', 'IVATasa16NoAcred', 'TotTasa11', 'BaseTasa11', 'IVATasa11', 'IVATasa11NoAcred'],
        ['Periodo de causación de IVA(R)', 'EjercicioAsignado', 'PeriodoAsignado'],
        ['Causación de IVA(C)', 'Tipo', 'TotTasa15', 'BaseTasa15', 'IVATasa15', 'TotTasa10', 'BaseTasa10', 'IVATasa10', 'TotTasa0', 'BaseTasa0', 'TotTasaExento', 'BaseTasaExento', 'TotOtraTasa', 'BaseOtraTasa', 'IVAOtraTasa', 'ISRRetenido', 'TotOtros', 'IVARetenido', 'Captado', 'NoCausar'],
        ['Asociación movimiento(AM)', 'UUID'],
        ['Movimiento de póliza(M1)', 'IdCuenta', 'Referencia', 'TipoMovto', 'Importe', 'IdDiario', 'ImporteME', 'Concepto', 'IdSegNeg', 'Guid'],
        ['Egreso(EG)', 'IdDocumentoDe', 'TipoDocumento', 'Folio', 'Fecha', 'FechaAplicacion', 'CodigoPersona', 'BeneficiarioPagador', 'IdCuentaCheques', 'CodigoMoneda', 'Total', 'Referencia', 'Origen', 'BancoDestino', 'CuentaDestino', 'OtroMetodoDePago', 'Guid'],
        ['Póliza(P)', 'Fecha', 'TipoPol', 'Folio', 'Clase', 'IdDiario', 'Concepto', 'SistOrig', 'Impresa', 'Ajuste', 'Guid'],
        ['Datos para Facturación Electrónica(FE)', 'RutaAnexo', 'ArchivoAnexo'],
        ['Cheque(CH)', 'IdDocumentoDe', 'TipoDocumento', 'Folio', 'Fecha', 'FechaAplicacion', 'CodigoPersona', 'BeneficiarioPagador', 'IdCuentaCheques', 'CodigoMoneda', 'Total', 'Referencia', 'Origen', 'CuentaDestino', 'BancoDestino', 'Guid'],
        ['Movimiento de póliza(M)', 'IdCuenta', 'Referencia', 'TipoMovto', 'Importe', 'IdDiario', 'ImporteME', 'Concepto', 'IdSegNeg'],
        ['Devolución de IVA (IETU)(W2)', 'IETUDeducible', 'IETUAcreditable', 'IETUModificado', 'IdConceptoIETU', ' '],
    ]
    fnt_enc = Font(name='Calibri', size=9, color='595959')
    for enc_row in ENCABEZADO_CONTPAQ:
        for c, val in enumerate(enc_row, 1):
            cell = ws_lay.cell(row=ws_lay.max_row + 1 if c == 1 else ws_lay.max_row,
                               column=c, value=val)
            cell.font = fnt_enc
        # avanzar fila solo al terminar cada fila de encabezado
    # — resetear y escribir de corrido (más limpio)
    ws_lay.delete_rows(1, ws_lay.max_row)
    for enc_row in ENCABEZADO_CONTPAQ:
        ws_lay.append(enc_row)
        for c in range(1, len(enc_row) + 1):
            ws_lay.cell(row=ws_lay.max_row, column=c).font = fnt_enc

    row_num = len(ENCABEZADO_CONTPAQ) + 1
    resumen = []

    for i, m in enumerate(todos_matches, 1):
        es_p2   = m.get('pasada', 1) == 2
        filas, avisos = generar_bloque(m, i, fact_map, catalogo)
        for fila in filas:
            tipo = fila[0]
            # Pasada 2: encabezado P en verde claro para distinguirlas visualmente
            fill = (fill_P2 if es_p2 else fill_P) if tipo == 'P' else (fill_AD if tipo == 'AD' else None)
            fnt  = fnt_P if tipo == 'P' else fnt_n
            for c, val in enumerate(fila, 1):
                cell = ws_lay.cell(row=row_num, column=c, value=val)
                if fill: cell.fill = fill
                cell.font = fnt
                cell.alignment = Alignment(vertical='center')
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'
                if tipo == 'P' and c == 2 and isinstance(val, datetime):
                    cell.number_format = 'yyyymmdd'
            row_num += 1

        grupo = m['grupo_rep']
        for fac_rep in grupo['facturas']:
            fd = fact_map.get(fac_rep['uuid_fac'].lower())
            resumen.append({
                'Póliza':           i,
                'Pasada':           m.get('pasada', 1),
                'Fecha banco':      m['banco_row'].get('Fecha_dt', ''),
                'Monto banco':      m['banco_row'].get('Importe', 0),
                'Cliente':          grupo['razon'],
                'RFC':              grupo['rfc'],
                'Folio':            fac_rep['folio'],
                'Total factura':    fd['Total_num']      if fd is not None else fac_rep['importe_fac'],
                'Importe cobrado':  fac_rep['importe_fac'],
                'IVA':              fd['IVA_num']        if fd is not None else 0,
                'IVA declarado':    'SI' if (fd is not None and fd['IVA_ya_declarado']) else '',
                'Método pago':      fd['Metodo_pago']    if fd is not None else '',
                'UUID Factura':     fac_rep['uuid_fac'],
                'UUID REP':         grupo['uuid_rep'] or '',
                'Cuenta CXC':       buscar_cuenta_cte(
                                        fd['RFC_limpio']      if fd is not None else grupo['rfc'],
                                        fd['Razon_receptor']  if fd is not None else grupo['razon'],
                                        catalogo),
                'Multi-folio':      'SI' if len(grupo['facturas']) > 1 else '',
                'En Fact_Pend':     'SI' if fd is not None else '⚠ NO',
                'Método match':     m.get('metodo', 'p1:rep'),
                'Avisos':           ' | '.join(avisos) if avisos else '',
            })

    anchos_lay = [5, 14, 10, 4, 14, 4, 4, 70, 4, 4]
    for i, w in enumerate(anchos_lay, 1):
        ws_lay.column_dimensions[get_column_letter(i)].width = w
    ws_lay.freeze_panes = 'A17'   # datos empiezan en fila 17 (tras 16 de encabezado)

    # ── Resumen Matches ───────────────────────────────────────────────────────
    ws_res = wb.create_sheet('Resumen Matches')
    if resumen:
        df_res = pd.DataFrame(resumen)
        hdr_row(ws_res, list(df_res.columns))
        for r, (_, row) in enumerate(df_res.iterrows(), 2):
            for c, val in enumerate(row, 1):
                cell = ws_res.cell(row=r, column=c, value=val)
                cell.font = fnt_n
                cell.alignment = Alignment(vertical='center')
                if isinstance(val, float):    cell.number_format = '#,##0.00'
                if isinstance(val, datetime): cell.number_format = 'DD/MM/YYYY'
                if df_res.columns[c-1] == 'En Fact_Pend' and val == '⚠ NO':
                    cell.fill = fill_warn
        for i in range(1, len(df_res.columns)+1):
            ws_res.column_dimensions[get_column_letter(i)].width = 22
        ws_res.freeze_panes = 'A2'

    # ── Sin Match — fallaron en ambas pasadas ────────────────────────────────
    ws_sm = wb.create_sheet('Sin Match - Revisar')
    hdrs_sm = ['Fecha', 'Importe', 'Nombre Ordenante', 'RFC Ordenante',
               'Concepto', 'Referencia', 'Acción sugerida']
    hdr_row(ws_sm, hdrs_sm)
    err_fill = PatternFill('solid', start_color='C00000')
    for c in range(1, 8):
        ws_sm.cell(1, c).fill = err_fill
    for r, row in enumerate(sin_match_p2, 2):
        vals = [
            row.get('Fecha_dt', ''), row.get('Importe', 0),
            row.get('Nombre Ordenante', ''), row.get('RFC Ordenante', ''),
            row.get('Concepto', ''), row.get('Referencia', ''),
            'Sin REP y sin match en Fact_Pendientes — verificar manualmente',
        ]
        for c, val in enumerate(vals, 1):
            cell = ws_sm.cell(row=r, column=c, value=val)
            cell.fill, cell.font = fill_sm, fnt_n
            if isinstance(val, float):    cell.number_format = '#,##0.00'
            if isinstance(val, datetime): cell.number_format = 'DD/MM/YYYY'
    for i in range(1, 8):
        ws_sm.column_dimensions[get_column_letter(i)].width = 26
    ws_sm.freeze_panes = 'A2'

    # ── Sin Match — REPs sin abono bancario ───────────────────────────────────
    if sin_match_rep:
        ws_smr = wb.create_sheet('REPs sin Abono - Revisar')
        hdrs_smr = ['UUID REP', 'Cliente', 'RFC', 'Fecha REP',
                    'Total REP', 'Folios', 'Acción sugerida']
        hdr_row(ws_smr, hdrs_smr)
        warn_fill = PatternFill('solid', start_color='7B1FA2')
        for c in range(1, 8):
            ws_smr.cell(1, c).fill = warn_fill
        for r, g in enumerate(sin_match_rep, 2):
            folios = ', '.join(f['folio'] for f in g['facturas'])
            vals = [
                g['uuid_rep'], g['razon'], g['rfc'],
                g['fecha_pago'].strftime('%d/%m/%Y') if g['fecha_pago'] else '',
                g['monto_dep'], folios,
                'Verificar si el depósito llegó en otro periodo',
            ]
            for c, val in enumerate(vals, 1):
                cell = ws_smr.cell(row=r, column=c, value=val)
                cell.fill = fill_smrep
                cell.font = fnt_n
                if isinstance(val, float): cell.number_format = '#,##0.00'
        for i in range(1, 8):
            ws_smr.column_dimensions[get_column_letter(i)].width = 28
        ws_smr.freeze_panes = 'A2'

    wb.save(ruta_out)

    return {
        'n_matches':     n_ok,
        'n_matches_p2':  n_ok_p2,
        'n_sin_match':   n_sm_p2,
        'n_multi':       n_multi,
        'n_sin_rep':     n_sm_p2,
        'n_rep_sin_ban': n_sm_rep,
        'n_iva_cero':    n_iva_cero,
        'tasa':          tasa_global,
        'monto_ok':      round(total_ok + total_ok_p2, 2),
        'monto_sm':      round(total_sm, 2),
    }


# ══════════════════════════════════════════════════════════════════════════════
# PUNTO DE ENTRADA
# ══════════════════════════════════════════════════════════════════════════════

import os
import subprocess

def convertir_a_xls(ruta_xlsx):
    """
    Convierte el .xlsx generado a .xls (MS Excel 97) usando LibreOffice headless.
    Retorna la ruta del .xls resultante, o None si falla.
    CONTPAq i solo reconoce .xls y solo lee la primera hoja.
    """
    outdir = os.path.dirname(ruta_xlsx)
    try:
        resultado = subprocess.run(
            ['soffice', '--headless', '--convert-to', 'xls', ruta_xlsx, '--outdir', outdir],
            capture_output=True, text=True, timeout=60
        )
        if resultado.returncode == 0:
            ruta_xls = ruta_xlsx.replace('.xlsx', '.xls')
            if os.path.exists(ruta_xls):
                os.remove(ruta_xlsx)   # limpiar el xlsx intermedio
                return ruta_xls
    except Exception:
        pass
    return None   # si falla, el xlsx sigue disponible como fallback


# ══════════════════════════════════════════════════════════════════════════════
# PUNTO DE ENTRADA
# ══════════════════════════════════════════════════════════════════════════════

def procesar_polizas(ruta_banco, ruta_facturas, ruta_catalogo, ruta_out,
                     ruta_reps=None):
    banco             = leer_banco(ruta_banco)
    fact_map, fact_df = leer_facturas(ruta_facturas)
    catalogo          = leer_catalogo(ruta_catalogo)
    grupos_rep        = leer_reps(ruta_reps) if ruta_reps else []

    # Pasada 1: REP como fuente primaria
    matches, sin_match_rep, sin_rep_banco = hacer_matching(grupos_rep, banco)

    # UUIDs ya usados en Pasada 1
    uuids_p1 = {
        fac['uuid_fac'].lower()
        for m in matches
        for fac in m['grupo_rep']['facturas']
    }

    # Pasada 2: depósitos sin REP vs Fact_Pendientes
    matches_p2, sin_match_p2 = matching_pasada2(sin_rep_banco, fact_df, uuids_p1)

    stats = escribir_excel(
        matches, sin_match_rep, sin_rep_banco,
        fact_map, catalogo, ruta_out,
        matches_p2=matches_p2,
        sin_match_p2=sin_match_p2,
    )

    # Convertir a .xls para carga directa en CONTPAq i
    ruta_xls = convertir_a_xls(ruta_out)
    stats['ruta_final'] = ruta_xls if ruta_xls else ruta_out

    return stats
